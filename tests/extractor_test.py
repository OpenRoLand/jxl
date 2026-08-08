"""Tests for :mod:`openroland_jxl.extractor` and the public convenience API."""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path

import openpyxl
from openroland_survey import database
from openroland_survey.models import SurveyPoint
from openroland_survey.records import IssueSeverity
from sqlalchemy.orm import Session

import openroland_jxl
from openroland_jxl.extractor import JxlExtractor


class TestJxlExtractorIdentity:
    """Tests for the extractor's protocol-facing identity attributes."""

    def test_format_name(self) -> None:
        assert JxlExtractor.format_name == "jxl"

    def test_extensions(self) -> None:
        assert JxlExtractor.extensions == (".jxl",)

    def test_can_read_matches_case_insensitively(self, tmp_path: Path) -> None:
        extractor = JxlExtractor()

        assert extractor.can_read(tmp_path / "job.JXL")
        assert extractor.can_read(tmp_path / "job.jxl")
        assert not extractor.can_read(tmp_path / "job.rw5")


class TestExtractUnqualifiedDocument:
    """Extraction of the minimal, no-namespace fixture."""

    def test_extracts_every_point(self, minimal_jxl: Path) -> None:
        result = JxlExtractor().extract(minimal_jxl)

        assert len(result.records) == 2

    def test_maps_canonical_fields(self, minimal_jxl: Path) -> None:
        result = JxlExtractor().extract(minimal_jxl)

        first = result.records[0]
        assert first.north == Decimal("500000.123")
        assert first.east == Decimal("300000.456")
        assert first.height == Decimal("120.789")
        assert first.name == "1"
        assert first.code == "TOP"
        assert first.description == "Boundary corner"
        assert first.method == "GPS"

    def test_source_metadata(self, minimal_jxl: Path) -> None:
        result = JxlExtractor().extract(minimal_jxl)

        assert result.source.source_type.value == "jxl"
        assert result.source.resolved_path == minimal_jxl.resolve()
        assert result.source.source_crs is None

    def test_has_no_issues(self, minimal_jxl: Path) -> None:
        result = JxlExtractor().extract(minimal_jxl)

        assert result.issues == ()


class TestExtractNamespacedDocument:
    """Extraction of the default-namespace fixture with field-book merge
    and Environment CRS metadata."""

    def test_merges_field_book_metadata(self, namespaced_jxl: Path) -> None:
        result = JxlExtractor().extract(namespaced_jxl)

        assert len(result.records) == 1
        record = result.records[0]
        assert record.code == "BM"
        assert record.description == "Benchmark near gate"

    def test_source_crs_from_environment(self, namespaced_jxl: Path) -> None:
        result = JxlExtractor().extract(namespaced_jxl)

        assert result.source.source_crs == "EPSG:3844"

    def test_source_record_id(self, namespaced_jxl: Path) -> None:
        result = JxlExtractor().extract(namespaced_jxl)

        assert result.records[0].source_record_id == "P100"


class TestExtractSurveyProDocument:
    """Extraction of the Survey Pro-style fixture with QC and WGS84."""

    def test_maps_wgs84_and_quality(self, survey_pro_jxl: Path) -> None:
        result = JxlExtractor().extract(survey_pro_jxl)

        record = next(r for r in result.records if r.name == "205")
        assert record.latitude == 46.65977287134
        assert record.longitude == 25.62071090333
        assert record.wgs84_altitude == 837.2458
        assert record.status == "FIXED"
        assert record.method == "GpsStaticObservation"
        assert record.hrms == 0.01
        assert record.vrms == 0.016
        assert record.pdop == 1.1
        assert record.hdop == 0.6
        assert record.vdop == 0.9
        assert record.satellite_count == 14
        assert record.observed_at_utc is not None
        assert record.source_record_id == "000000e1"
        assert record.source_values["TimeStamp"] == "2010-01-01T21:57:03"
        assert record.source_values["Method"] == "GpsStaticObservation"

    def test_deleted_field_book_record_not_merged(
        self, survey_pro_jxl: Path
    ) -> None:
        result = JxlExtractor().extract(survey_pro_jxl)

        names = [record.name for record in result.records]
        assert "999" in names
        deleted = next(r for r in result.records if r.name == "999")
        assert "TimeStamp" not in deleted.source_values


class TestExtractKeyedInAndMeasured:
    """Extraction of the mixed KeyedIn/measured/invalid fixture."""

    def test_excludes_keyed_in_by_default(
        self, keyed_in_and_measured_jxl: Path
    ) -> None:
        result = JxlExtractor().extract(keyed_in_and_measured_jxl)

        names = [record.name for record in result.records]
        assert "10" not in names
        assert "11" in names

    def test_includes_keyed_in_when_requested(
        self, keyed_in_and_measured_jxl: Path
    ) -> None:
        extractor = JxlExtractor(include_keyed_in=True)
        result = extractor.extract(keyed_in_and_measured_jxl)

        names = [record.name for record in result.records]
        assert "10" in names

    def test_skips_non_numeric_coordinate_with_warning(
        self, keyed_in_and_measured_jxl: Path
    ) -> None:
        result = JxlExtractor().extract(keyed_in_and_measured_jxl)

        names = [record.name for record in result.records]
        assert "12" not in names
        warnings = [
            issue
            for issue in result.issues
            if issue.severity == IssueSeverity.WARNING
        ]
        assert len(warnings) == 1
        assert warnings[0].record_id == "12"


class TestClassifySurveyMethod:
    """Measured quality tokens and coord-only Code remap on extract."""

    def _write_point(self, path: Path, method: str) -> None:
        path.write_bytes(
            (
                '<?xml version="1.0"?>\n'
                "<JOBFile>\n"
                "  <Reductions>\n"
                "    <Point>\n"
                "      <Name>1</Name>\n"
                "      <SurveyMethod>%s</SurveyMethod>\n"
                "      <Grid>\n"
                "        <North>1</North>\n"
                "        <East>2</East>\n"
                "        <Elevation>3</Elevation>\n"
                "      </Grid>\n"
                "    </Point>\n"
                "  </Reductions>\n"
                "</JOBFile>\n" % method
            ).encode("utf-8")
        )

    def test_fix_becomes_fixed_status(self, tmp_path: Path) -> None:
        path = tmp_path / "fix.jxl"
        self._write_point(path, "Fix")
        record = JxlExtractor().extract(path).records[0]
        assert record.status == "FIXED"
        assert record.method is None
        assert record.kind is None
        assert record.source_values.get("SurveyMethod") == "Fix"

    def test_network_float_becomes_float_status(self, tmp_path: Path) -> None:
        path = tmp_path / "nfloat.jxl"
        self._write_point(path, "NetworkFloat")
        record = JxlExtractor().extract(path).records[0]
        assert record.status == "FLOAT"
        assert record.method is None

    def test_code_becomes_imported(self, tmp_path: Path) -> None:
        path = tmp_path / "code.jxl"
        self._write_point(path, "Code")
        record = JxlExtractor().extract(path).records[0]
        assert record.method is None
        assert record.status is None
        assert record.kind == "imported"

    def test_base_survey_method_becomes_kind_base(self, tmp_path: Path) -> None:
        path = tmp_path / "base.jxl"
        self._write_point(path, "Base")
        record = JxlExtractor().extract(path).records[0]
        assert record.kind == "base"
        assert record.method is None
        assert record.status is None
        assert record.source_values.get("SurveyMethod") == "Base"

    def test_base_station_survey_method_becomes_kind_base(
        self, tmp_path: Path
    ) -> None:
        path = tmp_path / "base_station.jxl"
        self._write_point(path, "Base Station")
        record = JxlExtractor().extract(path).records[0]
        assert record.kind == "base"
        assert record.method is None

    def test_gps_method_unchanged(self, tmp_path: Path) -> None:
        path = tmp_path / "gps.jxl"
        self._write_point(path, "GPS")
        record = JxlExtractor().extract(path).records[0]
        assert record.method == "GPS"
        assert record.status is None
        assert record.kind is None


class TestExtractPreservesFieldBookRawValues:
    """Field-book values with no canonical column land in ``source_values``."""

    def test_field_book_extras_are_nested(self, tmp_path: Path) -> None:
        source_path = tmp_path / "with_operator.jxl"
        source_path.write_bytes(b"""<?xml version="1.0"?>
<JOBFile>
  <FieldBook>
    <PointRecord>
      <ID>50</ID>
      <Operator>Jane</Operator>
    </PointRecord>
  </FieldBook>
  <Reductions>
    <Point>
      <ID>50</ID>
      <Grid>
        <North>1</North>
        <East>2</East>
        <Elevation>3</Elevation>
      </Grid>
    </Point>
  </Reductions>
</JOBFile>
""")

        result = JxlExtractor().extract(source_path)

        assert result.records[0].source_values["field_book"] == {
            "Operator": "Jane"
        }


class TestExtractMissingOrNonFiniteCoordinates:
    """A point with missing or non-finite Grid coordinates is skipped."""

    def test_missing_grid_produces_warning(self, tmp_path: Path) -> None:
        source_path = tmp_path / "no_grid.jxl"
        source_path.write_bytes(b"""<?xml version="1.0"?>
<JOBFile>
  <Reductions>
    <Point>
      <Name>30</Name>
    </Point>
  </Reductions>
</JOBFile>
""")

        result = JxlExtractor().extract(source_path)

        assert result.records == ()
        assert len(result.issues) == 1
        assert result.issues[0].severity == IssueSeverity.WARNING

    def test_non_finite_grid_value_produces_warning(
        self, tmp_path: Path
    ) -> None:
        source_path = tmp_path / "nan_grid.jxl"
        source_path.write_bytes(b"""<?xml version="1.0"?>
<JOBFile>
  <Reductions>
    <Point>
      <Name>31</Name>
      <Grid>
        <North>1</North>
        <East>NaN</East>
        <Elevation>3</Elevation>
      </Grid>
    </Point>
  </Reductions>
</JOBFile>
""")

        result = JxlExtractor().extract(source_path)

        assert result.records == ()
        assert len(result.issues) == 1
        assert result.issues[0].severity == IssueSeverity.WARNING


class TestExtractPointsConvenienceFunction:
    """Tests for the top-level :func:`openroland_jxl.extract_points`."""

    def test_default_excludes_keyed_in(
        self, keyed_in_and_measured_jxl: Path
    ) -> None:
        result = openroland_jxl.extract_points(keyed_in_and_measured_jxl)

        names = [record.name for record in result.records]
        assert "10" not in names

    def test_include_keyed_in_option(
        self, keyed_in_and_measured_jxl: Path
    ) -> None:
        result = openroland_jxl.extract_points(
            keyed_in_and_measured_jxl, include_keyed_in=True
        )

        names = [record.name for record in result.records]
        assert "10" in names


class TestExportToXlsxConvenienceFunction:
    """Tests for the top-level :func:`openroland_jxl.export_to_xlsx`."""

    def test_writes_expected_row_count(
        self, minimal_jxl: Path, tmp_path: Path
    ) -> None:
        output_path = tmp_path / "out.xlsx"

        summary = openroland_jxl.export_to_xlsx(minimal_jxl, output_path)

        assert summary.written_row_count == 2
        assert output_path.exists()

    def test_workbook_contains_points_sheet(
        self, minimal_jxl: Path, tmp_path: Path
    ) -> None:
        output_path = tmp_path / "out.xlsx"
        openroland_jxl.export_to_xlsx(minimal_jxl, output_path)

        workbook = openpyxl.load_workbook(output_path)
        try:
            assert "Points" in workbook.sheetnames
            assert "Source" in workbook.sheetnames
        finally:
            workbook.close()


class TestExportToDatabaseConvenienceFunction:
    """Tests for the top-level :func:`openroland_jxl.export_to_database`."""

    def test_imports_expected_point_count(
        self, minimal_jxl: Path, tmp_path: Path
    ) -> None:
        db_path = tmp_path / "survey.sqlite3"

        summary = openroland_jxl.export_to_database(
            minimal_jxl, db_path, source_crs="EPSG:3844"
        )

        assert summary.unique_points_inserted == 2
        assert summary.source_count == 1

    def test_points_are_queryable_afterwards(
        self, namespaced_jxl: Path, tmp_path: Path
    ) -> None:
        db_path = tmp_path / "survey.sqlite3"

        openroland_jxl.export_to_database(namespaced_jxl, db_path)

        engine = database.create_engine(db_path)
        try:
            with Session(engine) as session:
                points = session.query(SurveyPoint).all()
                assert len(points) == 1
                assert points[0].code == "BM"
        finally:
            engine.dispose()

    def test_accepts_an_existing_engine(
        self, minimal_jxl: Path, tmp_path: Path
    ) -> None:
        db_path = tmp_path / "survey.sqlite3"
        engine = database.create_engine(db_path)
        try:
            summary = openroland_jxl.export_to_database(
                minimal_jxl, engine, source_crs="EPSG:3844"
            )
            assert summary.unique_points_inserted == 2
        finally:
            engine.dispose()
